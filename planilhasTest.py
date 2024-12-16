#
#Codigo Desenvolvido por Nicolas Martins Lorena
#
from openpyxl import Workbook, load_workbook
import pandas as pd
import json
import unicodedata
from difflib import get_close_matches


class ConverterPlanilhas:

    def __init__(self, entrada_planilha):
        # Carregar a planilha existente
        self.workbook = load_workbook(entrada_planilha)
        self.sheet = self.workbook.active  # Selecionar a primeira planilha

    def normalizar_texto(self, texto):
        """Normaliza texto para evitar problemas com acentos e padroniza para minúsculas."""
        return unicodedata.normalize('NFKD', str(texto)).encode('ascii', 'ignore').decode('utf-8').strip().lower()

    def gera_dicionario(self, preco_planilha, saida_json):
        """Gera um dicionário JSON a partir da planilha de preços."""
        df = pd.read_excel(preco_planilha, sheet_name=0)

        # Converte os dados para um dicionário
        dados = df.to_dict(orient='records')

        # Filtra os dados
        dados_filtrados = [
            {
                'ID': item['ID'],
                'Identificação': self.normalizar_texto(item['Identificacao']),
                'Exame/Serviço': item['Exame/Serviço']
            }
            for item in dados
        ]

        # Salva o dicionário como um arquivo JSON
        with open(saida_json, 'w', encoding='utf-8') as json_file:
            json.dump(dados_filtrados, json_file, ensure_ascii=False, indent=4)

        print(f"Dicionário salvo em {saida_json}")

    def buscar_por_partes(self, identificacao, identificacoes_disponiveis):
        """Tenta encontrar identificadores no dicionário usando partes do nome."""
        partes = identificacao.split()
        for parte in partes:
            matches = [id for id in identificacoes_disponiveis if parte in id]
            if matches:
                return matches[0]
        return None

    def obter_informacoes(self, identificacao_procurado, arquivo_json):
        """Busca informações no dicionário JSON com base no Identificador."""
        identificacao_procurado_normalizado = self.normalizar_texto(identificacao_procurado)
        with open(arquivo_json, 'r', encoding='utf-8') as f:
            dados = json.load(f)

            # Criar lista de identificações normalizadas
            identificacoes_disponiveis = [item['Identificação'] for item in dados]

            # Tenta encontrar uma correspondência exata
            if identificacao_procurado_normalizado in identificacoes_disponiveis:
                return next(item for item in dados if item['Identificação'] == identificacao_procurado_normalizado)

            # Busca identificadores similares (usando difflib)
            similar = get_close_matches(
                identificacao_procurado_normalizado,
                identificacoes_disponiveis,
                n=1,
                cutoff=0.5  # Define a similaridade mínima (0.0 a 1.0)
            )

            if similar:
                return next(item for item in dados if item['Identificação'] == similar[0])

            # Tenta correspondência parcial por palavras-chave
            parcial = self.buscar_por_partes(identificacao_procurado_normalizado, identificacoes_disponiveis)
            if parcial:
                return next(item for item in dados if item['Identificação'] == parcial)

        print(f"Identificador '{identificacao_procurado}' não encontrado no dicionário.")
        return None

    def processar_planilhas(self, arquivo_json, nova_planilha):
        """Processa a planilha de entrada e gera uma nova planilha com as informações dos identificadores e valores."""
    
        try:
            workbook = load_workbook(nova_planilha)
            print(f"Planilha '{nova_planilha}' carregada com sucesso.")
        except FileNotFoundError:
            workbook = Workbook()
            print(f"Planilha '{nova_planilha}' criada.")

        nova_sheet = workbook.active
        nova_sheet.title = "Dados Atualizados"

        # Adiciona cabeçalho na nova planilha (sempre na primeira linha)
        if nova_sheet.max_row == 1 and nova_sheet.cell(1, 1).value is None:
            nova_sheet.append(['ID', 'Identificacao', 'Exame/Serviço', 'Valor Interno', 'Valor In Company'])
        else:
            nova_sheet.delete_rows(1)  # Remove qualquer linha extra existente no início
            nova_sheet.append(['ID', 'Identificacao', 'Exame/Serviço', 'Valor Interno', 'Valor In Company'])

        # Itera sobre a planilha original
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, min_col=3, max_col=4, values_only=True):
            identificador, valor = row
            if identificador:
                informacao = self.obter_informacoes(identificador, arquivo_json)

                # Verifica se encontrou as informações
                if informacao:
                    # Verifica se o valor está completo e válido (não nulo ou vazio)
                    if valor is not None and valor != '':
                        nova_sheet.append([ 
                            informacao['ID'],
                            identificador,
                            informacao['Exame/Serviço'],
                            valor,
                        ])
                    else:
                        print(f"Erro: Valor Interno não encontrado para o identificador '{identificador}'. Não será adicionado à planilha.")
                else:
                    print(f"Erro: Identificador '{identificador}' não encontrado no dicionário. Não será adicionado à planilha.")

        workbook.save(nova_planilha)
        print(f"Dados processados e salvos em '{nova_planilha}' com sucesso.")


# Exemplo de uso:
entrada_planilha = 'doismilturismo.xlsx'
preco_planilha = 'tabela.xlsx'
saida_json = 'dicionarioFinal.json'
nova_planilha = 'nova_planilha.xlsx'

# Gera dicionário apenas se não existir
teste = ConverterPlanilhas(entrada_planilha)
teste.gera_dicionario(preco_planilha, saida_json)
teste.processar_planilhas(saida_json, nova_planilha)



