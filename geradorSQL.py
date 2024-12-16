#
# Codigo Desenvolvido por Nicolas Martins Lorena
#
from openpyxl import load_workbook
import json
import unicodedata
from difflib import get_close_matches


class ConverterPlanilhas:

    def __init__(self, entrada_planilha):
        # Carregar a planilha existente
        self.workbook = load_workbook(entrada_planilha)
        self.sheet = self.workbook.active  # Selecionar a primeira planilha

    def normalizar_texto(self, texto):
        """Normaliza texto para evitar problemas com acentos e padroniza para minúsculas."""
        return unicodedata.normalize('NFKD', str(texto)).encode('ascii', 'ignore').decode('utf-8').strip().lower()

    def buscar_por_partes(self, identificacao, identificacoes_disponiveis):
        """Tenta encontrar identificadores no dicionário usando partes do nome."""
        partes = identificacao.split()
        for parte in partes:
            matches = [id for id in identificacoes_disponiveis if parte in id]
            if matches:
                return matches[0]
        return None

    def obter_informacoes(self, identificacao_procurado, arquivo_json):
        """Busca informações no dicionário JSON com base no Identificador."""
        identificacao_procurado_normalizado = self.normalizar_texto(identificacao_procurado)
        with open(arquivo_json, 'r', encoding='utf-8') as f:
            dados = json.load(f)

            # Criar lista de identificações normalizadas
            identificacoes_disponiveis = [item['Identificação'] for item in dados]

            # Tenta encontrar uma correspondência exata
            if identificacao_procurado_normalizado in identificacoes_disponiveis:
                return next(item for item in dados if item['Identificação'] == identificacao_procurado_normalizado)

            # Busca identificadores similares (usando difflib)
            similar = get_close_matches(
                identificacao_procurado_normalizado,
                identificacoes_disponiveis,
                n=1,
                cutoff=0.6  # Define a similaridade mínima (0.0 a 1.0)
            )

            if similar:
                return next(item for item in dados if item['Identificação'] == similar[0])

            # Tenta correspondência parcial por palavras-chave
            parcial = self.buscar_por_partes(identificacao_procurado_normalizado, identificacoes_disponiveis)
            if parcial:
                return next(item for item in dados if item['Identificação'] == parcial)

        print(f"Identificador '{identificacao_procurado}' não encontrado no dicionário.")
        return None

    def gerar_sql(self, arquivo_json, arquivo_sql):
        """Gera um arquivo SQL com base nas informações da planilha e do dicionário JSON."""
        sql_statements = []

        # Adiciona cabeçalho para o script SQL
        sql_statements.append("-- Script SQL gerado automaticamente\n")
        sql_statements.append("-- Inserindo dados na tabela 'exames_servicos'\n")

        # Itera sobre a planilha original
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, min_col=3, max_col=4, values_only=True):
            identificador, valor = row
            if identificador:
                informacao = self.obter_informacoes(identificador, arquivo_json)

                # Verifica se encontrou as informações
                if informacao:
                    if valor is not None and valor != '':
                        # Monta o comando SQL
                        sql = (
                            f"INSERT INTO exames_servicos (id, identificacao, exame_servico, valor_interno, valor_in_company) "
                            f"VALUES ({informacao['ID']}, '{identificador}', '{informacao['Exame/Serviço']}', {valor}, {0});"
                        )
                        sql_statements.append(sql)
                    else:
                        print(f"Erro: Valor Interno não encontrado para o identificador '{identificador}'. Não será adicionado ao SQL.")
                else:
                    print(f"Erro: Identificador '{identificador}' não encontrado no dicionário. Não será adicionado ao SQL.")

        # Salva os comandos SQL no arquivo
        with open(arquivo_sql, 'w', encoding='utf-8') as f:
            f.write("\n".join(sql_statements))

        print(f"Script SQL gerado e salvo em '{arquivo_sql}' com sucesso.")


# Exemplo de uso:
entrada_planilha = 'doismilturismo.xlsx'
saida_json = 'dicionarioFinal.json'  # Use o dicionário já existente
arquivo_sql = 'inserir_exames.sql'

# Apenas gera o arquivo SQL com o dicionário existente
teste = ConverterPlanilhas(entrada_planilha)
teste.gerar_sql(saida_json, arquivo_sql)
